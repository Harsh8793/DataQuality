"""Report service: generate PDF/Excel/JSON/CSV export artifacts."""

from __future__ import annotations

import json

from sqlalchemy.orm import Session

from app.core.storage import get_storage
from app.exceptions.base import NotFoundException
from app.models.report import GeneratedReport
from app.repositories.analysis_repository import QualityIssueRepository, QualityReportRepository
from app.repositories.dataset_repository import DatasetColumnRepository
from app.repositories.report_repository import ReportRepository
from app.schemas.chat import ReportResponse
from app.services.base import BaseService, DatasetContextMixin


class ReportService(BaseService, DatasetContextMixin):
    """Builds downloadable reports from a dataset's latest quality report."""

    def __init__(self, db: Session) -> None:
        super().__init__(db)
        self.storage = get_storage()
        self.reports = ReportRepository(db)
        self.quality_reports = QualityReportRepository(db)
        self.issues = QualityIssueRepository(db)
        self.columns = DatasetColumnRepository(db)

    def generate(self, dataset_id: str, user_id: str, report_type: str) -> ReportResponse:
        """Generate a report of the requested type and persist its metadata."""
        dataset = self._load_owned_dataset(dataset_id, user_id)
        report = self.quality_reports.latest_for_dataset(dataset_id)
        if report is None:
            raise NotFoundException("Analyze the dataset before generating a report.")
        issues = self.issues.list_for_report(report.id)

        filename = f"{dataset_id}_{report_type}.{report_type}"
        path = self.storage.report_path(filename)
        payload = self._build_payload(dataset, report, issues)
        payload["profile"] = self._build_profile(dataset_id)
        payload["dashboard"] = self._build_dashboard(dataset_id, user_id)

        if report_type == "pdf":
            self._write_pdf(path, payload)
        elif report_type == "xlsx":
            self._write_xlsx(path, payload)
        elif report_type == "csv":
            self._write_csv(path, payload)
        else:
            path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")

        record = self.reports.create(
            user_id=user_id, dataset_id=dataset_id, report_type=report_type,
            title=f"{dataset.name} - Quality Report", file_path=str(path),
            size_bytes=path.stat().st_size, created_by=user_id,
        )
        self.db.commit()
        return ReportResponse.model_validate(record)

    def get_record(self, report_id: str, user_id: str) -> GeneratedReport:
        """Return a report record owned by the user, for download."""
        record = self.reports.get(report_id)
        if record is None or record.user_id != user_id:
            raise NotFoundException("Report not found.")
        return record

    # ---- profile + dashboard summaries ------------------------------- #
    def _build_profile(self, dataset_id) -> list[dict]:
        """Column profile (key summary): type, completeness, distinctness, PII."""
        return [
            {
                "name": c.name, "type": c.semantic_type,
                "null_pct": c.null_pct, "distinct": c.distinct_count,
                "pii": bool(c.is_pii),
            }
            for c in self.columns.list_for_dataset(dataset_id)
        ]

    def _build_dashboard(self, dataset_id, user_id) -> dict:
        """The user's selected dashboard KPIs + charts (default if not customized)."""
        from app.services.dashboard_service import DashboardService

        try:
            builder = DashboardService(self.db).get_builder(dataset_id, user_id)
        except Exception:  # noqa: BLE001 - export must not fail if dashboard can't build
            return {"kpis": [], "charts": []}
        kpi_ids = set(builder.selected.kpis)
        chart_ids = set(builder.selected.charts)
        return {
            "kpis": [
                {"label": k.label, "value": k.value, "format": k.format}
                for k in builder.pool.kpis if k.id in kpi_ids
            ],
            "charts": [
                {"title": c.title, "type": c.type, "x": c.x, "y": c.y, "data": c.data}
                for c in builder.pool.charts if c.id in chart_ids
            ],
        }

    # ---- payload + writers ------------------------------------------- #
    def _build_payload(self, dataset, report, issues) -> dict:
        return {
            "dataset": {"id": dataset.id, "name": dataset.name, "rows": dataset.row_count, "columns": dataset.col_count},
            "scores": {
                "overall": report.overall_score, "completeness": report.completeness,
                "accuracy": report.accuracy, "consistency": report.consistency,
                "uniqueness": report.uniqueness, "validity": report.validity, "integrity": report.integrity,
            },
            "summary": {"total_issues": report.total_issues, "duplicate_rows": report.duplicate_rows},
            "issues": [
                {"check": i.check_key, "column": i.column_name, "severity": i.severity,
                 "count": i.count, "fix": i.recommended_fix}
                for i in issues
            ],
        }

    def _write_pdf(self, path, payload) -> None:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        styles = getSampleStyleSheet()
        doc = SimpleDocTemplate(str(path), pagesize=A4, title="DataPilot AI Report")
        ds = payload["dataset"]
        story = [
            Paragraph("DataPilot AI - Data Quality Report", styles["Title"]),
            Paragraph(f"Dataset: {ds['name']}", styles["Normal"]),
            Paragraph(f"{ds['rows']:,} rows &middot; {ds['columns']} columns", styles["Normal"]),
            Spacer(1, 12),
            Paragraph(f"Overall Score: {payload['scores']['overall']}/100", styles["Heading2"]),
        ]
        score_rows = [["Dimension", "Score"]] + [[k.title(), f"{v}"] for k, v in payload["scores"].items() if k != "overall"]
        score_table = Table(score_rows, hAlign="LEFT")
        score_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story += [score_table, Spacer(1, 16)]

        header_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#374151")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ])

        # ---- Key metrics (dashboard KPIs) ----
        dash = payload.get("dashboard", {})
        if dash.get("kpis"):
            story += [Paragraph("Key Metrics", styles["Heading2"])]
            kpi_rows = [["Metric", "Value"]] + [
                [k["label"], self._fmt_value(k["value"], k.get("format"))] for k in dash["kpis"]
            ]
            t = Table(kpi_rows, hAlign="LEFT")
            t.setStyle(header_style)
            story += [t, Spacer(1, 16)]

        # ---- Dataset profile (key summary) ----
        if payload.get("profile"):
            story += [Paragraph("Dataset Profile", styles["Heading2"])]
            prof_rows = [["Column", "Type", "Null %", "Distinct", "PII"]]
            for c in payload["profile"][:60]:
                prof_rows.append([c["name"], c["type"], f"{c['null_pct']}%", str(c["distinct"]), "Yes" if c["pii"] else "-"])
            t = Table(prof_rows, hAlign="LEFT")
            t.setStyle(header_style)
            story += [t, Spacer(1, 16)]

        # ---- Dashboard charts (rendered as real vector charts) ----
        if dash.get("charts"):
            story += [Paragraph("Dashboard", styles["Heading2"])]
            for ch in dash["charts"]:
                story += [Paragraph(f"{ch['title']} ({ch['type']})", styles["Heading4"])]
                drawing = self._pdf_chart(ch)
                if drawing is not None:
                    story += [drawing, Spacer(1, 12)]
                else:
                    # Fallback to a data table if the chart can't be drawn.
                    data = ch.get("data", [])[:12]
                    if data and "name" in data[0]:
                        rows = [["Category", "Value"]] + [[str(d.get("name")), str(d.get("value"))] for d in data]
                    elif data:
                        rows = [[ch.get("x", "x"), ch.get("y", "y")]] + [[str(d.get("x")), str(d.get("y"))] for d in data]
                    else:
                        rows = [["No data"]]
                    t = Table(rows, hAlign="LEFT")
                    t.setStyle(header_style)
                    story += [t, Spacer(1, 12)]

        # ---- Top issues ----
        story += [Paragraph("Top Issues", styles["Heading2"])]
        issue_rows = [["Check", "Column", "Severity", "Count"]]
        for i in payload["issues"][:25]:
            issue_rows.append([i["check"], i["column"] or "-", i["severity"], str(i["count"])])
        issue_table = Table(issue_rows, hAlign="LEFT")
        issue_table.setStyle(header_style)
        story.append(issue_table)
        doc.build(story)

    @staticmethod
    def _fmt_value(value, fmt) -> str:
        """Human-format a KPI value for reports."""
        try:
            num = float(value)
        except (TypeError, ValueError):
            return str(value)
        text = f"{num:,.0f}" if num.is_integer() else f"{num:,.2f}"
        return f"${text}" if fmt == "currency" else text

    # Chart palette (matches the dark UI accents).
    _PALETTE = ("#3b82f6", "#8b5cf6", "#06b6d4", "#22c55e", "#f59e0b", "#ef4444", "#ec4899", "#14b8a6")

    def _pdf_chart(self, ch):
        """Draw a real vector chart (bar/pie/line/scatter) with reportlab.graphics."""
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.graphics.charts.lineplots import LinePlot
        from reportlab.graphics.charts.piecharts import Pie
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.widgets.markers import makeMarker
        from reportlab.lib import colors

        data = ch.get("data", [])[:12]
        if not data:
            return None
        palette = [colors.HexColor(c) for c in self._PALETTE]
        ctype = ch.get("type")

        try:
            if ctype == "scatter":
                pts = [
                    (float(d["x"]), float(d["y"]))
                    for d in data if d.get("x") is not None and d.get("y") is not None
                ]
                if not pts:
                    return None
                dr = Drawing(430, 210)
                lp = LinePlot()
                lp.x, lp.y, lp.width, lp.height = 45, 30, 350, 160
                lp.data = [pts]
                lp.lines[0].strokeColor = colors.transparent
                lp.lines[0].strokeWidth = 0
                lp.lines[0].symbol = makeMarker("Circle")
                lp.lines[0].symbol.fillColor = palette[1]
                lp.lines[0].symbol.size = 5
                dr.add(lp)
                return dr

            names = [str(d.get("name", ""))[:18] for d in data]
            values = [float(d.get("value") or 0) for d in data]

            if ctype == "pie":
                dr = Drawing(430, 210)
                pie = Pie()
                pie.x, pie.y, pie.width, pie.height = 150, 20, 170, 170
                pie.data = values
                pie.labels = names
                pie.sideLabels = True
                for i in range(len(values)):
                    pie.slices[i].fillColor = palette[i % len(palette)]
                dr.add(pie)
                return dr

            if ctype == "line":
                from reportlab.graphics.charts.linecharts import HorizontalLineChart

                dr = Drawing(430, 210)
                lc = HorizontalLineChart()
                lc.x, lc.y, lc.width, lc.height = 45, 45, 350, 150
                lc.data = [values]
                lc.categoryAxis.categoryNames = names
                lc.categoryAxis.labels.angle = 30
                lc.categoryAxis.labels.boxAnchor = "ne"
                lc.lines[0].strokeColor = palette[0]
                lc.lines[0].strokeWidth = 2
                dr.add(lc)
                return dr

            # Default: bar chart.
            dr = Drawing(430, 210)
            bc = VerticalBarChart()
            bc.x, bc.y, bc.width, bc.height = 45, 45, 350, 150
            bc.data = [values]
            bc.categoryAxis.categoryNames = names
            bc.categoryAxis.labels.angle = 30
            bc.categoryAxis.labels.boxAnchor = "ne"
            bc.categoryAxis.labels.dy = -4
            bc.valueAxis.valueMin = 0
            bc.bars[0].fillColor = palette[0]
            dr.add(bc)
            return dr
        except Exception:  # noqa: BLE001 - fall back to a data table on any drawing error
            return None

    def _write_xlsx(self, path, payload) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Scores"
        ws.append(["Dataset", payload["dataset"]["name"]])
        ws.append(["Rows", payload["dataset"]["rows"]])
        ws.append(["Columns", payload["dataset"]["columns"]])
        ws.append([])
        ws.append(["Dimension", "Score"])
        for k, v in payload["scores"].items():
            ws.append([k.title(), v])

        # Dashboard: key metrics on one sheet; real charts on another.
        dash = payload.get("dashboard", {})
        wsd = wb.create_sheet("Dashboard")
        wsd.append(["Key Metrics"])
        wsd.append(["Metric", "Value"])
        for k in dash.get("kpis", []):
            wsd.append([k["label"], self._fmt_value(k["value"], k.get("format"))])
        if dash.get("charts"):
            self._write_xlsx_charts(wb, dash["charts"])

        # Column profile (key summary).
        wsp = wb.create_sheet("Profile")
        wsp.append(["Column", "Type", "Null %", "Distinct", "PII"])
        for c in payload.get("profile", []):
            wsp.append([c["name"], c["type"], c["null_pct"], c["distinct"], "Yes" if c["pii"] else ""])

        ws2 = wb.create_sheet("Issues")
        ws2.append(["Check", "Column", "Severity", "Count", "Recommended Fix"])
        for i in payload["issues"]:
            ws2.append([i["check"], i["column"], i["severity"], i["count"], i["fix"]])
        wb.save(str(path))

    def _write_xlsx_charts(self, wb, charts) -> None:
        """Render real Excel charts (bar/pie/line/scatter) from chart specs.

        Data goes in hidden left columns; each chart is anchored to the right,
        stacked vertically, so the sheet shows graphs (not just tables).
        """
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart, Series

        ws = wb.create_sheet("Charts")
        data_row = 1  # running row cursor in the (left) data columns
        anchor_row = 1  # running row for placing charts on the right

        for ch in charts:
            data = ch.get("data", [])[:12]
            if not data:
                continue
            ctype = ch.get("type")
            title_row = data_row
            ws.cell(row=title_row, column=1, value=ch.get("title", "Chart"))
            data_row += 1

            if ctype == "scatter":
                header = data_row
                ws.cell(row=header, column=1, value=str(ch.get("x", "x")))
                ws.cell(row=header, column=2, value=str(ch.get("y", "y")))
                first = header + 1
                for d in data:
                    ws.cell(row=data_row + 1, column=1, value=d.get("x"))
                    ws.cell(row=data_row + 1, column=2, value=d.get("y"))
                    data_row += 1
                last = data_row
                chart = ScatterChart()
                chart.x_axis.title = str(ch.get("x", "x"))
                chart.y_axis.title = str(ch.get("y", "y"))
                xref = Reference(ws, min_col=1, min_row=first, max_row=last)
                yref = Reference(ws, min_col=2, min_row=first, max_row=last)
                series = Series(yref, xref, title=ch.get("title", "Chart"))
                series.marker.symbol = "circle"
                series.graphicalProperties.line.noFill = True
                chart.series.append(series)
            else:
                header = data_row
                ws.cell(row=header, column=1, value="Category")
                ws.cell(row=header, column=2, value=ch.get("title", "Value"))
                first = header + 1
                for d in data:
                    ws.cell(row=data_row + 1, column=1, value=str(d.get("name")))
                    ws.cell(row=data_row + 1, column=2, value=d.get("value"))
                    data_row += 1
                last = data_row
                chart = PieChart() if ctype == "pie" else (LineChart() if ctype == "line" else BarChart())
                cats = Reference(ws, min_col=1, min_row=first, max_row=last)
                vals = Reference(ws, min_col=2, min_row=header, max_row=last)  # include header for legend title
                chart.add_data(vals, titles_from_data=True)
                chart.set_categories(cats)

            chart.title = ch.get("title", "Chart")
            chart.height = 7
            chart.width = 14
            ws.add_chart(chart, f"E{anchor_row}")
            anchor_row += 15
            data_row += 2  # blank gap before the next chart's data block

    def _write_csv(self, path, payload) -> None:
        import csv

        with path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(["Check", "Column", "Severity", "Count", "Recommended Fix"])
            for i in payload["issues"]:
                writer.writerow([i["check"], i["column"], i["severity"], i["count"], i["fix"]])
